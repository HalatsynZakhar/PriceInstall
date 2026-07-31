from __future__ import annotations

import io
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from horoshop_prices import ArticleMappings, CatalogIndex, FieldChange, PricePlan, PriceRow, WholesaleChange, build_excel_template, build_failed_excel, load_article_mappings, load_settings, normalize_article_code, parse_article_mappings, parse_excel_prices, plan_prices, save_article_mappings


class HoroshopPricesTests(unittest.TestCase):
    def settings(self):
        with tempfile.TemporaryDirectory() as directory:
            config = Path(directory) / "config.json"
            config.write_text('{"horoshop":{"domain":"https://shop.test"},"wholesale_defaults":{"recalculate_from_rrc":true,"rules":[{"tier":1,"minimal_threshold":2,"discount_percent":10},{"tier":2,"minimal_threshold":5,"discount_percent":15},{"tier":3,"minimal_threshold":10,"discount_percent":25}]}}', encoding="utf-8")
            return load_settings(config)

    def excel(self, headers, rows):
        workbook = Workbook()
        workbook.active.append(headers)
        for row in rows:
            workbook.active.append(row)
        output = io.BytesIO(); workbook.save(output); workbook.close()
        return output.getvalue()

    def catalog(self):
        return CatalogIndex.from_raw([{"article":"REAL-1","article_for_display":"Display-1","price":1000,"price_old":1000,"wholesale_prices":[{"minimal_threshold":2,"price":900},{"minimal_threshold":20,"price":700}]}])

    def multi_catalog(self):
        return CatalogIndex.from_raw([
            {"article":"REAL-1","article_for_display":"Display-1","price":1000,"price_old":1000,"wholesale_prices":[]},
            {"article":"REAL-2","article_for_display":"Display-2","price":1100,"price_old":1100,"wholesale_prices":[]},
        ])

    def test_sale_price_disables_wholesale_price_above_it(self):
        row = PriceRow("Display-1", FieldChange(Decimal("750")), None, FieldChange(), False, False, (), 2)
        plan = plan_prices([row], self.catalog(), self.settings())[0]
        self.assertEqual(plan.payload["wholesale_prices"], [{"minimal_threshold": 20, "price": 700.0}])
        self.assertIn("Вимкнено", "; ".join(plan.warnings))

    def test_moves_existing_current_price_to_rrc(self):
        row = PriceRow("REAL-1", FieldChange(Decimal("800")), None, FieldChange(), True, False, (), 2)
        plan = plan_prices([row], self.catalog(), self.settings())[0]
        self.assertEqual(plan.payload["price"], 800.0)
        self.assertEqual(plan.payload["price_old"], 1000.0)

    def test_explicit_delete_is_different_from_blank(self):
        row = PriceRow("REAL-1", FieldChange(), None, FieldChange(delete=True), False, False, (WholesaleChange(1, FieldChange(delete=True)),), 2)
        plan = plan_prices([row], self.catalog(), self.settings())[0]
        self.assertEqual(plan.payload["price_old"], 0)
        self.assertEqual(plan.payload["wholesale_prices"], [{"minimal_threshold": 20, "price": 700.0}])

    def test_moves_rrc_to_current_price(self):
        row = PriceRow("REAL-1", FieldChange(), None, FieldChange(), False, True, (), 2)
        plan = plan_prices([row], self.catalog(), self.settings())[0]
        self.assertEqual(plan.payload["price"], 1000.0)

    def test_current_price_can_be_calculated_as_fractional_percent_of_rrc(self):
        row = PriceRow("REAL-1", FieldChange(), Decimal("75.6"), FieldChange(), False, False, (), 2)
        plan = plan_prices([row], self.catalog(), self.settings())[0]
        self.assertEqual(plan.payload["price"], 756.0)

    def test_excel_supports_one_row_and_reordered_columns(self):
        rows = parse_excel_prices(self.excel(["Опт 1", "Артикул", "Поточна ціна", "Поточну ціну в РРЦ (Так)", "Опт 1 від"], [(800, "A", 900, "Так", 3)]))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].wholesale[0].threshold, 3)
        self.assertTrue(rows[0].move_current_to_old)

    def test_empty_cells_do_not_create_operation(self):
        with self.assertRaisesRegex(ValueError, "немає жодної команди"):
            parse_excel_prices(self.excel(["Артикул", "Поточна ціна"], [("A", "")]))

    def test_missing_required_article_header_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "обов'язковий стовпець"):
            parse_excel_prices(self.excel(["Поточна ціна"], [(100,)]))

    def test_duplicate_service_headers_are_rejected(self):
        with self.assertRaisesRegex(ValueError, "повторюється"):
            parse_excel_prices(self.excel(["Артикул", "Опт 1", "Опт 1"], [("A", 100, 90)]))

    def test_template_contains_all_price_columns(self):
        workbook = load_workbook(io.BytesIO(build_excel_template()), read_only=True)
        headers = [cell.value for cell in next(workbook["Ціни"].iter_rows(max_row=1))]
        self.assertIn("Поточну ціну в РРЦ (Так)", headers)
        self.assertIn("Поточна ціна (% від РРЦ)", headers)
        self.assertIn("Опт 5 від (шт.)", headers)
        workbook.close()

    def test_normalized_article_lookup_is_optional(self):
        row = PriceRow("Display.1", FieldChange(Decimal("900")), None, FieldChange(), False, False, (), 2)
        without_normalization = plan_prices([row], self.catalog(), self.settings())[0]
        with_normalization = plan_prices([row], self.catalog(), self.settings(), normalize_articles=True)[0]
        self.assertFalse(without_normalization.ready)
        self.assertEqual(with_normalization.article, "REAL-1")
        self.assertEqual(normalize_article_code("A  (B)__C...D"), "A-B-C-D")

    def test_mapping_file_can_expand_one_source_article_to_multiple_targets(self):
        mappings = parse_article_mappings(self.excel(["Артикул", "Подартикул"], [("SET-1", "REAL-1"), ("SET-1", "REAL-2")]))
        row = PriceRow("SET-1", FieldChange(Decimal("900")), None, FieldChange(), False, False, (), 2)
        plans = plan_prices([row], self.multi_catalog(), self.settings(), mappings=mappings)
        self.assertEqual([plan.article for plan in plans], ["REAL-1", "REAL-2"])
        self.assertEqual([plan.payload["price"] for plan in plans], [900.0, 900.0])

    def test_mapping_database_merges_and_persists_rules(self):
        initial = ArticleMappings({"SET-1": ("REAL-1",)})
        additional = ArticleMappings({"SET-1": ("REAL-2",), "SET-2": ("REAL-3",)})
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "data" / "article_mappings.json"
            save_article_mappings(path, initial.merged_with(additional))
            restored = load_article_mappings(path)
        self.assertEqual(restored.entries, {"SET-1": ("REAL-1", "REAL-2"), "SET-2": ("REAL-3",)})

    def test_failed_excel_contains_user_price_fields(self):
        row = PriceRow("UNKNOWN", FieldChange(Decimal("900")), None, FieldChange(delete=True), False, False, (WholesaleChange(1, FieldChange(Decimal("800")), 3),), 7)
        data = build_failed_excel([(PricePlan("", "UNKNOWN", {}, (row,), error="not found"), "Помилка", "not found")])
        workbook = load_workbook(io.BytesIO(data), read_only=True)
        values = next(workbook["Не встановлено"].iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertEqual(values[1], "not found")
        self.assertEqual(values[4], 7)
        self.assertEqual(values[5], "UNKNOWN")
        self.assertEqual(values[6], 900)
        self.assertEqual(values[8], "Видалити")
        self.assertEqual(values[11], 800)
        self.assertEqual(values[12], 3)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
