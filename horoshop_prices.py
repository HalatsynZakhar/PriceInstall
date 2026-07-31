from __future__ import annotations

import io
import json
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_DOWN, ROUND_HALF_UP
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_EXPORT_LIMIT = 500
PUBLIC_LOG_PATH_JSON_PATTERN = re.compile(r'("public_log_path"\s*:\s*")(?P<value>[^"]*)(")')
PRICE_TYPES = {
    "поточна ціна": "price", "ціна": "price", "price": "price", "current": "price", "акційна ціна": "price",
    "ррц": "price_old", "стара ціна": "price_old", "old price": "price_old", "price_old": "price_old",
    "опт 1": "wholesale_1", "оптова ціна 1": "wholesale_1", "wholesale 1": "wholesale_1",
    "опт 2": "wholesale_2", "оптова ціна 2": "wholesale_2", "wholesale 2": "wholesale_2",
    "опт 3": "wholesale_3", "оптова ціна 3": "wholesale_3", "wholesale 3": "wholesale_3",
    "опт 4": "wholesale_4", "оптова ціна 4": "wholesale_4", "wholesale 4": "wholesale_4",
    "опт 5": "wholesale_5", "оптова ціна 5": "wholesale_5", "wholesale 5": "wholesale_5",
}


class HoroshopPricesError(RuntimeError):
    pass


@dataclass(frozen=True)
class WholesaleRule:
    tier: int
    minimal_threshold: int
    discount_percent: Decimal


@dataclass(frozen=True)
class Settings:
    domain: str
    host: str
    port: int
    batch_size: int
    request_timeout_seconds: int
    public_log_path: Path
    public_log_name: str
    wholesale_rules: tuple[WholesaleRule, ...]
    recalculate_wholesale_from_rrc: bool

    @property
    def public_log_file(self) -> Path:
        return self.public_log_path / self.public_log_name


@dataclass(frozen=True)
class Credentials:
    login: str = ""
    password: str = ""
    token: str = ""


@dataclass(frozen=True)
class FieldChange:
    value: Decimal | None = None
    delete: bool = False

    @property
    def specified(self) -> bool:
        return self.value is not None or self.delete


@dataclass(frozen=True)
class WholesaleChange:
    tier: int
    change: FieldChange
    threshold: int | None = None


@dataclass(frozen=True)
class PriceRow:
    display_article: str
    current: FieldChange
    current_percent_of_old: Decimal | None
    old: FieldChange
    move_current_to_old: bool
    move_old_to_current: bool
    wholesale: tuple[WholesaleChange, ...]
    row_number: int


@dataclass(frozen=True)
class CatalogProduct:
    article: str
    display_article: str
    price: Decimal | None
    price_old: Decimal | None
    wholesale_prices: tuple[dict[str, Any], ...]


@dataclass(frozen=True)
class PricePlan:
    article: str
    display_article: str
    payload: dict[str, Any]
    rows: tuple[PriceRow, ...]
    warnings: tuple[str, ...] = ()
    error: str = ""

    @property
    def ready(self) -> bool:
        return not self.error


def normalize(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_article_code(value: Any) -> str:
    text = normalize(value)
    if not text:
        return ""
    text = "".join(character if character.isalnum() else "-" for character in text)
    text = re.sub(r"-{2,}", "-", text)
    return text.strip("-")


def endpoint_url(domain: str, endpoint: str) -> str:
    return urljoin(f"{domain.rstrip('/')}/", endpoint.lstrip("/"))


def parse_decimal(value: Any, field: str = "ціна", allow_zero: bool = False) -> Decimal:
    text = normalize(value).replace(" ", "").replace(",", ".")
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError) as error:
        raise ValueError(f"{field} має бути числом, більшим за нуль.") from error
    if number < 0 or (number == 0 and not allow_zero):
        raise ValueError(f"{field} має бути більшим за нуль.")
    return number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def parse_threshold(value: Any) -> int | None:
    text = normalize(value)
    if not text:
        return None
    try:
        number = int(text)
    except ValueError as error:
        raise ValueError("кількість від має бути цілим числом.") from error
    if number < 1:
        raise ValueError("кількість від має бути не меншою за 1.")
    return number


def parse_price_type(value: Any) -> str:
    key = normalize(value).casefold()
    price_type = PRICE_TYPES.get(key)
    if not price_type:
        raise ValueError("невідомий тип ціни. Використайте: Поточна ціна, РРЦ або Опт 1-5.")
    return price_type


def _repair_path(config_text: str) -> str:
    def replace(match: re.Match[str]) -> str:
        value = re.sub(r"(?<!\\)\\(?!\\)", lambda _: "\\\\", match.group("value"))
        return f"{match.group(1)}{value}{match.group(3)}"
    return PUBLIC_LOG_PATH_JSON_PATTERN.sub(replace, config_text)


def load_settings(config_file: Path) -> Settings:
    config_text = config_file.read_text(encoding="utf-8-sig")
    try:
        raw = json.loads(config_text)
    except json.JSONDecodeError as original_error:
        repaired = _repair_path(config_text)
        if repaired == config_text:
            raise original_error
        raw = json.loads(repaired)
        config_file.write_text(repaired, encoding="utf-8")
    if not isinstance(raw, dict):
        raise ValueError("config.json має містити об'єкт.")
    server = raw.get("server") or {}
    horoshop = raw.get("horoshop") or {}
    logging_config = raw.get("logging") or {}
    prices = raw.get("wholesale_defaults") or {}
    domain = normalize(horoshop.get("domain"))
    if not domain:
        raise ValueError("Вкажіть horoshop.domain у config.json.")
    rules: list[WholesaleRule] = []
    raw_rules = prices.get("rules") if isinstance(prices, dict) else None
    if raw_rules is None:
        raw_rules = [
            {"tier": 1, "minimal_threshold": 2, "discount_percent": 10},
            {"tier": 2, "minimal_threshold": 5, "discount_percent": 15},
            {"tier": 3, "minimal_threshold": 10, "discount_percent": 25},
        ]
    if not isinstance(raw_rules, list):
        raise ValueError("wholesale_defaults.rules має бути списком.")
    for raw_rule in raw_rules:
        if not isinstance(raw_rule, dict):
            raise ValueError("Кожне правило оптової ціни має бути об'єктом.")
        tier = int(raw_rule.get("tier", 0))
        threshold = int(raw_rule.get("minimal_threshold", 0))
        discount = parse_decimal(raw_rule.get("discount_percent"), "відсоток знижки")
        if tier not in range(1, 6) or threshold < 1 or discount >= 100:
            raise ValueError("Правила опту повинні містити рівень 1-5, поріг від 1 та знижку до 100%.")
        rules.append(WholesaleRule(tier, threshold, discount))
    path = Path(normalize(logging_config.get("public_log_path", "logs")) or "logs")
    if not path.is_absolute():
        path = config_file.parent / path
    name = normalize(logging_config.get("public_log_name", "horoshop_prices.log")) or "horoshop_prices.log"
    if Path(name).name != name:
        raise ValueError("logging.public_log_name має бути лише назвою файлу.")
    return Settings(
        domain=domain.rstrip("/"), host=normalize(server.get("host", "0.0.0.0")) or "0.0.0.0",
        port=max(1, min(65535, int(server.get("port", 8095)))), batch_size=max(1, int(horoshop.get("batch_size", 50))),
        request_timeout_seconds=max(1, int(horoshop.get("request_timeout_seconds", 60))), public_log_path=path, public_log_name=name,
        wholesale_rules=tuple(sorted(rules, key=lambda rule: rule.tier)),
        recalculate_wholesale_from_rrc=bool(prices.get("recalculate_from_rrc", prices.get("recalculate_on_current_price", True))) if isinstance(prices, dict) else True,
    )


DELETE_MARKERS = {"видалити", "delete", "так", "yes"}
YES_MARKERS = {"так", "да", "yes", "1"}
HEADER_ARTICLE = {"артикул", "артикул для відображення", "article", "article_for_display"}
@dataclass(frozen=True)
class ArticleMappings:
    entries: dict[str, tuple[str, ...]]

    @classmethod
    def empty(cls) -> "ArticleMappings":
        return cls({})

    def merged_with(self, other: "ArticleMappings") -> "ArticleMappings":
        combined = {source: list(targets) for source, targets in self.entries.items()}
        for source, targets in other.entries.items():
            combined.setdefault(source, []).extend(targets)
        return ArticleMappings({source: tuple(dict.fromkeys(targets)) for source, targets in combined.items()})

    def targets_for(self, article: str, use_normalized: bool) -> tuple[str, ...]:
        def other_variants(targets: tuple[str, ...]) -> tuple[str, ...]:
            variants = tuple(target for target in targets if normalize(target).casefold() != article.casefold())
            return variants or (article,)

        if not self.entries:
            return (article,)
        exact = self.entries.get(article)
        if exact:
            return other_variants(exact)
        folded = article.casefold()
        folded_matches = [targets for source, targets in self.entries.items() if source.casefold() == folded]
        if len(folded_matches) == 1:
            return other_variants(folded_matches[0])
        if use_normalized:
            normalized = normalize_article_code(article).casefold()
            normalized_matches = [targets for source, targets in self.entries.items() if normalize_article_code(source).casefold() == normalized]
            if len(normalized_matches) == 1:
                return other_variants(normalized_matches[0])
        return (article,)


def header_key(value: Any) -> str:
    return " ".join(normalize(value).casefold().split())


def parse_change(value: Any, label: str) -> FieldChange:
    text = normalize(value)
    if not text:
        return FieldChange()
    if text.casefold() in DELETE_MARKERS:
        return FieldChange(delete=True)
    return FieldChange(value=parse_decimal(text, label))


def parse_yes(value: Any, label: str) -> bool:
    text = normalize(value)
    if not text or text.casefold() in {"ні", "нет", "no", "0"}:
        return False
    if text.casefold() in YES_MARKERS:
        return True
    raise ValueError(f"у колонці «{label}» вкажіть Так або залиште порожньою.")


def column_map(headers: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    aliases = {"поточна ціна": "price", "поточна ціна (грн)": "price", "поточна ціна, % від ррц": "percent", "поточна ціна (% від ррц)": "percent", "ррц": "old", "ррц (грн)": "old", "поточну ціну в ррц (так)": "move", "ррц в поточну ціну (так)": "restore"}
    for tier in range(1, 6):
        aliases[f"опт {tier}"] = f"wholesale_{tier}"
        aliases[f"опт {tier} (грн)"] = f"wholesale_{tier}"
        aliases[f"опт {tier} від"] = f"threshold_{tier}"
        aliases[f"опт {tier} від (шт.)"] = f"threshold_{tier}"
    for index, value in enumerate(headers):
        key = header_key(value)
        target = "article" if key in HEADER_ARTICLE else aliases.get(key)
        if target:
            if target in mapping:
                raise ValueError(f"Заголовок «{value}» повторюється.")
            mapping[target] = index
    if "article" not in mapping:
        raise ValueError("Не знайдено обов'язковий стовпець «Артикул».")
    return mapping


def cell(row: tuple[Any, ...], mapping: dict[str, int], key: str) -> Any:
    index = mapping.get(key)
    return row[index] if index is not None and index < len(row) else None


def parse_excel_prices(data: bytes) -> list[PriceRow]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = next(iterator)
        except StopIteration as error:
            raise ValueError("Excel-файл порожній.") from error
        mapping = column_map(headers)
        rows: list[PriceRow] = []
        for row_number, row in enumerate(iterator, start=2):
            if not row or all(value is None for value in row):
                continue
            try:
                article = normalize(cell(row, mapping, "article"))
                if not article:
                    raise ValueError("вкажіть артикул.")
                current = parse_change(cell(row, mapping, "price"), "Поточна ціна")
                if current.delete:
                    raise ValueError("поточну ціну не можна видалити. Вкажіть нове значення.")
                percent_text = normalize(cell(row, mapping, "percent"))
                percent = parse_decimal(percent_text, "Відсоток від РРЦ") if percent_text else None
                if percent is not None and not Decimal("1") <= percent <= Decimal("100"):
                    raise ValueError("відсоток від РРЦ має бути від 1 до 100.")
                if current.value is not None and percent is not None:
                    raise ValueError("вкажіть або поточну ціну, або відсоток від РРЦ, але не обидва значення.")
                old = parse_change(cell(row, mapping, "old"), "РРЦ")
                move = parse_yes(cell(row, mapping, "move"), "Поточну ціну в РРЦ (Так)")
                restore = parse_yes(cell(row, mapping, "restore"), "РРЦ в поточну ціну (Так)")
                if move and restore:
                    raise ValueError("в одному рядку можна обрати лише один автоматичний перенос ціни.")
                if move and old.specified:
                    raise ValueError("не поєднуйте РРЦ та «Поточну ціну в РРЦ (Так)» в одному рядку.")
                if move and current.value is None and percent is None:
                    raise ValueError("для перенесення поточної ціни в РРЦ вкажіть нову поточну ціну або відсоток від РРЦ.")
                if restore and old.value is not None:
                    raise ValueError("для перенесення РРЦ у поточну ціну не вказуйте нове значення РРЦ.")
                if restore and percent is not None:
                    raise ValueError("не поєднуйте перенесення РРЦ у поточну ціну з відсотком від РРЦ.")
                if percent is not None and old.delete:
                    raise ValueError("не можна видаляти РРЦ, з якого обчислюється поточна ціна.")
                wholesale: list[WholesaleChange] = []
                for tier in range(1, 6):
                    change = parse_change(cell(row, mapping, f"wholesale_{tier}"), f"Опт {tier}")
                    threshold = parse_threshold(cell(row, mapping, f"threshold_{tier}"))
                    if threshold is not None and not change.specified:
                        raise ValueError(f"для Опт {tier} задано кількість без ціни або команди «Видалити».")
                    if change.specified:
                        wholesale.append(WholesaleChange(tier, change, threshold))
                if not current.specified and percent is None and not old.specified and not move and not restore and not wholesale:
                    continue
                rows.append(PriceRow(article, current, percent, old, move, restore, tuple(wholesale), row_number))
            except ValueError as error:
                raise ValueError(f"Рядок {row_number}: {error}") from error
        if not rows:
            raise ValueError("У файлі немає жодної команди. Порожні клітинки не змінюють ціни.")
        return rows
    finally:
        workbook.close()


def mapping_column_indexes(headers: tuple[Any, ...]) -> tuple[int, ...]:
    indexes = tuple(index for index, value in enumerate(headers) if normalize(value))
    if len(indexes) < 2:
        raise ValueError("У файлі сопоставлень має бути мінімум два стовпці з варіантами артикула.")
    return indexes


def parse_article_mappings(data: bytes) -> ArticleMappings:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = next(iterator)
        except StopIteration as error:
            raise ValueError("Файл сопоставлень порожній.") from error
        column_indexes = mapping_column_indexes(headers)
        mappings: dict[str, list[str]] = {}
        for row in iterator:
            if not row or all(value is None for value in row):
                continue
            articles = list(dict.fromkeys(
                normalize(row[index] if index < len(row) else None)
                for index in column_indexes
                if normalize(row[index] if index < len(row) else None)
            ))
            if not articles:
                continue
            if len(articles) < 2:
                continue
            for source in articles:
                mappings.setdefault(source, []).extend(articles)
        if not mappings:
            raise ValueError("У файлі сопоставлень немає жодного правила.")
        return ArticleMappings({source: tuple(dict.fromkeys(targets)) for source, targets in mappings.items()})
    finally:
        workbook.close()


def load_article_mappings(path: Path) -> ArticleMappings:
    if not path.exists():
        return ArticleMappings.empty()
    try:
        raw_entries = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise HoroshopPricesError(f"Не вдалося прочитати базу сопоставлень: {path}.") from error
    if not isinstance(raw_entries, dict):
        raise HoroshopPricesError("База сопоставлень має некоректний формат.")
    entries: dict[str, tuple[str, ...]] = {}
    for source, targets in raw_entries.items():
        source_article = normalize(source)
        if not source_article or not isinstance(targets, list):
            raise HoroshopPricesError("База сопоставлень має некоректний формат.")
        target_articles = tuple(dict.fromkeys(normalize(target) for target in targets if normalize(target)))
        if not target_articles:
            raise HoroshopPricesError("База сопоставлень має некоректний формат.")
        entries[source_article] = target_articles
    return ArticleMappings(entries)


def save_article_mappings(path: Path, mappings: ArticleMappings) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    serialized = {source: list(targets) for source, targets in mappings.entries.items()}
    temporary_path = path.with_suffix(path.suffix + ".tmp")
    try:
        temporary_path.write_text(json.dumps(serialized, ensure_ascii=False, indent=2), encoding="utf-8")
        temporary_path.replace(path)
    except OSError as error:
        raise HoroshopPricesError(f"Не вдалося зберегти базу сопоставлень: {path}.") from error


def price_row_excel_values(row: PriceRow) -> list[Any]:
    values: list[Any] = [
        row.row_number,
        row.display_article,
        float(row.current.value) if row.current.value is not None else "",
        float(row.current_percent_of_old) if row.current_percent_of_old is not None else "",
        "Видалити" if row.old.delete else (float(row.old.value) if row.old.value is not None else ""),
        "Так" if row.move_current_to_old else "",
        "Так" if row.move_old_to_current else "",
    ]
    by_tier = {item.tier: item for item in row.wholesale}
    for tier in range(1, 6):
        item = by_tier.get(tier)
        if item is None:
            values.extend(["", ""])
        else:
            values.extend(["Видалити" if item.change.delete else float(item.change.value or 0), item.threshold or ""])
    return values


def build_failed_excel(failures: list[tuple[PricePlan, str, str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Не встановлено"
    headers = ["Статус", "Помилка", "Внутрішній артикул Хорошоп", "Артикул для відображення", "Рядок", "Артикул користувача", "Поточна ціна (грн)", "Поточна ціна (% від РРЦ)", "РРЦ (грн)", "Поточну ціну в РРЦ (Так)", "РРЦ в поточну ціну (Так)"]
    for tier in range(1, 6):
        headers.extend([f"Опт {tier} (грн)", f"Опт {tier} від (шт.)"])
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(row=1, column=len(headers)).coordinate}"
    for plan, status, message in failures:
        row = plan.rows[0]
        sheet.append([status, message, plan.article, plan.display_article] + price_row_excel_values(row))
    widths = {
        "A": 14, "B": 56, "C": 26, "D": 26, "E": 10, "F": 28, "G": 18, "H": 24, "I": 14, "J": 26, "K": 26,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for index in range(12, len(headers) + 1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = 14
    _style_header(sheet)
    for row in sheet.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _style_header(worksheet: Any) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6D3B8F")
        cell.alignment = Alignment(horizontal="center")


def build_excel_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ціни"
    headers = ["Артикул", "Поточна ціна (грн)", "Поточна ціна (% від РРЦ)", "РРЦ (грн)", "Поточну ціну в РРЦ (Так)", "РРЦ в поточну ціну (Так)"]
    for tier in range(1, 6):
        headers.extend([f"Опт {tier} (грн)", f"Опт {tier} від (шт.)"])
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:P1"
    for column, width in {"A": 28, "B": 16, "C": 24, "D": 16, "E": 28, "F": 28, "G": 14, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14, "N": 14, "O": 14, "P": 14}.items():
        sheet.column_dimensions[column].width = width
    _style_header(sheet)
    examples = [("01063", "", 75, "", "Так", "", "", "", "", "", "", "", 650, 20, "", "")]
    for row in examples:
        sheet.append(row)
    for row in range(2, 202):
        sheet.cell(row=row, column=1).number_format = "@"
    guide = workbook.create_sheet("Інструкція")
    guide.column_dimensions["A"].width = 110
    guide["A1"] = "Масове встановлення цін Хорошоп"
    guide["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor="6D3B8F")
    guide.merge_cells("A1:B1")
    notes = [
        "Один рядок = один товар. Кожен стовпець керує окремою ціною цього товару.",
        "Артикул можна вказати як article_for_display зі сайту або внутрішній article Хорошоп. Пошук: точний, потім без урахування регістру.",
        "Порожня клітинка означає «не змінювати». Видалення виконується тільки словом «Видалити» у клітинці РРЦ або потрібного Опт 1-5.",
        "«Поточна ціна, % від РРЦ»: вкажіть число від 1 до 100, у тому числі дробове. Поточна ціна = РРЦ × відсоток / 100 з округленням до цілої гривні вниз.",
        "«Поточну ціну в РРЦ (Так)»: коли вказано нову поточну ціну, попередня поточна ціна буде записана в РРЦ. Не заповнюйте при цьому стовпець РРЦ.",
        "«РРЦ в поточну ціну (Так)» копіює наявне РРЦ у поточну ціну. Щоб одночасно прибрати РРЦ, напишіть «Видалити» у стовпці РРЦ.",
        "Опт 1-3 розраховується від РРЦ/старої ціни за правилами 2 шт. -10%, 5 шт. -15%, 10 шт. -25%. Правила можна змінити у config.json.",
        "Акційна ціна не зменшує опт. Оптові ціни, які не нижчі за поточну ціну, автоматично виключаються. Інші наявні рівні опта зберігаються.",
        "Для будь-якого Опт 1-5 можна вказати «Кількість від». Якщо для Опт 1-3 вона порожня, поріг береться з config.json. Для Опт 4-5 кількість обов'язкова.",
        "Стовпці можна переставляти. Якщо необов'язковий стовпець видалено, він не буде змінюватися. Стовпець «Артикул» обов'язковий.",
    ]
    for index, note in enumerate(notes, start=3):
        guide.cell(index, 1, note).alignment = Alignment(wrap_text=True, vertical="top")
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class CatalogIndex:
    def __init__(self, products: list[CatalogProduct]) -> None:
        self.products = products
        self.by_article = {item.article: item for item in products}
        self.by_exact: dict[str, list[CatalogProduct]] = {}
        self.by_folded: dict[str, list[CatalogProduct]] = {}
        self.by_normalized: dict[str, list[CatalogProduct]] = {}
        for item in products:
            for key in {item.article, item.display_article} - {""}:
                self.by_exact.setdefault(key, []).append(item)
                self.by_folded.setdefault(key.casefold(), []).append(item)
                normalized = normalize_article_code(key)
                if normalized:
                    self.by_normalized.setdefault(normalized.casefold(), []).append(item)

    @classmethod
    def from_raw(cls, raw_products: list[dict[str, Any]]) -> "CatalogIndex":
        products: list[CatalogProduct] = []
        for raw in raw_products:
            article = normalize(raw.get("article")) if isinstance(raw, dict) else ""
            if not article:
                continue
            wholesale = raw.get("wholesale_prices", [])
            products.append(CatalogProduct(
                article, normalize(raw.get("article_for_display")), decimal_or_none(raw.get("price")), decimal_or_none(raw.get("price_old")),
                tuple(item for item in wholesale if isinstance(item, dict)) if isinstance(wholesale, list) else (),
            ))
        return cls(products)

    def resolve(self, value: str, use_normalized: bool = False) -> tuple[CatalogProduct | None, str]:
        exact = list({item.article: item for item in self.by_exact.get(value, [])}.values())
        if len(exact) == 1:
            return exact[0], ""
        if len(exact) > 1:
            return None, f"Артикул '{value}' не є унікальним."
        folded = list({item.article: item for item in self.by_folded.get(value.casefold(), [])}.values())
        if len(folded) == 1:
            return folded[0], ""
        if len(folded) > 1:
            return None, f"Артикул '{value}' не є унікальним."
        if use_normalized:
            normalized = normalize_article_code(value)
            normalized_matches = list({item.article: item for item in self.by_normalized.get(normalized.casefold(), [])}.values())
            if len(normalized_matches) == 1:
                return normalized_matches[0], ""
            if len(normalized_matches) > 1:
                return None, f"Нормалізований артикул '{normalized}' для '{value}' не є унікальним."
        return None, f"Артикул '{value}' не знайдений у каталозі."


def decimal_or_none(value: Any) -> Decimal | None:
    text = normalize(value)
    if not text or text in {"0", "0.0", "0.00"}:
        return None
    try:
        return Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return None


def rule_for(settings: Settings, tier: int) -> WholesaleRule | None:
    return next((rule for rule in settings.wholesale_rules if rule.tier == tier), None)


def wholesale_entries(existing: tuple[dict[str, Any], ...]) -> dict[int, Decimal]:
    entries: dict[int, Decimal] = {}
    for item in existing:
        try:
            threshold = int(item.get("minimal_threshold"))
            price = parse_decimal(item.get("price"))
        except (TypeError, ValueError):
            continue
        entries[threshold] = price
    return entries


def plan_prices(
    rows: list[PriceRow],
    catalog: CatalogIndex,
    settings: Settings,
    normalize_articles: bool = False,
    mappings: ArticleMappings | None = None,
) -> list[PricePlan]:
    grouped: dict[str, PriceRow] = {}
    grouped_products: dict[str, CatalogProduct] = {}
    errors: list[PricePlan] = []
    mappings = mappings or ArticleMappings.empty()
    for row in rows:
        products: dict[str, CatalogProduct] = {}
        resolution_errors: list[str] = []
        for target_article in mappings.targets_for(row.display_article, normalize_articles):
            product, error = catalog.resolve(target_article, normalize_articles)
            if product is not None:
                products[product.article] = product
            elif error and "не знайдений у каталозі" not in error:
                resolution_errors.append(error)
        if not products:
            error = resolution_errors[0] if resolution_errors else f"Жоден варіант артикула для '{row.display_article}' не знайдений у каталозі."
            errors.append(PricePlan("", row.display_article, {}, (row,), error=error))
            continue
        for product in products.values():
            if product.article in grouped:
                errors.append(PricePlan(product.article, product.display_article or product.article, {}, (row,), error="Артикул повторюється у файлі або базі сопоставлень. Для товару має бути лише одна команда."))
            else:
                grouped[product.article] = row
                grouped_products[product.article] = product
    plans = errors
    for article, row in grouped.items():
        product = grouped_products[article]
        current = product.price
        old = product.price_old
        has_current = row.current.value is not None or row.current_percent_of_old is not None or row.move_old_to_current
        has_old = row.old.specified or row.move_current_to_old
        direct_wholesale = bool(row.wholesale)
        if row.move_old_to_current:
            current = product.price_old
        elif row.current_percent_of_old is not None:
            if old is None:
                plans.append(PricePlan(article, product.display_article or article, {}, (row,), error="Немає РРЦ для обчислення поточної ціни за відсотком."))
                continue
            current = (old * row.current_percent_of_old / Decimal("100")).quantize(Decimal("1"), rounding=ROUND_DOWN)
        elif row.current.value is not None:
            current = row.current.value
        if row.move_current_to_old:
            old = product.price
        elif row.old.delete:
            old = None
        elif row.old.value is not None:
            old = row.old.value
        if row.move_old_to_current and product.price_old is None:
            plans.append(PricePlan(article, product.display_article or article, {}, (row,), error="Немає РРЦ, яке можна перенести в поточну ціну."))
            continue
        if current is None:
            plans.append(PricePlan(article, product.display_article or article, {}, (row,), error="Немає поточної ціни: спочатку встановіть «Поточна ціна«."))
            continue
        wholesale = wholesale_entries(product.wholesale_prices)
        warnings: list[str] = []
        if (has_current or has_old) and settings.recalculate_wholesale_from_rrc and old is not None:
            for rule in settings.wholesale_rules:
                if rule.tier <= 3:
                    wholesale[rule.minimal_threshold] = (old * (Decimal("100") - rule.discount_percent) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            warnings.append("Опт 1-3 розраховано від РРЦ за правилами за замовчуванням.")
        for wholesale_change in row.wholesale:
            tier = wholesale_change.tier
            rule = rule_for(settings, tier)
            threshold = wholesale_change.threshold or (rule.minimal_threshold if rule else None)
            if threshold is None:
                plans.append(PricePlan(article, product.display_article or article, {}, (row,), error=f"Для Опт {tier} вкажіть «Кількість від»."))
                break
            if rule is not None and threshold != rule.minimal_threshold:
                wholesale.pop(rule.minimal_threshold, None)
            if wholesale_change.change.delete:
                wholesale.pop(threshold, None)
                warnings.append(f"Видалено Опт {tier} від {threshold} шт.")
            else:
                wholesale[threshold] = wholesale_change.change.value  # type: ignore[index]
        else:
            invalid = [threshold for threshold, value in wholesale.items() if value >= current]
            for threshold in invalid:
                del wholesale[threshold]
            if invalid:
                warnings.append("Вимкнено оптові рівні, де ціна не нижча за поточну: " + ", ".join(map(str, sorted(invalid))) + ".")
            payload: dict[str, Any] = {"article": article}
            if has_current: payload["price"] = float(current)
            if has_old: payload["price_old"] = float(old) if old else 0
            if has_current or direct_wholesale:
                payload["wholesale_prices"] = [{"minimal_threshold": threshold, "price": float(value)} for threshold, value in sorted(wholesale.items())]
            plans.append(PricePlan(article, product.display_article or article, payload, (row,), tuple(warnings)))
    return sorted(plans, key=lambda plan: min(row.row_number for row in plan.rows))


class HoroshopClient:
    def __init__(self, settings: Settings, credentials: Credentials) -> None:
        self.settings, self.credentials, self.session, self._token = settings, credentials, requests.Session(), credentials.token

    def _post(self, endpoint: str, payload: dict[str, Any]) -> dict[str, Any]:
        try:
            response = self.session.post(endpoint_url(self.settings.domain, endpoint), json=payload, timeout=self.settings.request_timeout_seconds)
            response.raise_for_status()
            data = response.json()
        except requests.RequestException as error:
            raise HoroshopPricesError(f"Помилка запиту до Хорошоп: {error}") from error
        except ValueError as error:
            raise HoroshopPricesError("Хорошоп повернув відповідь не у форматі JSON.") from error
        if not isinstance(data, dict):
            raise HoroshopPricesError("Хорошоп повернув некоректну відповідь.")
        if str(data.get("status", "")).upper() in {"ERROR", "EXCEPTION"}:
            raise HoroshopPricesError(str(data))
        return data

    def token(self) -> str:
        if self._token: return self._token
        if not self.credentials.login or not self.credentials.password:
            raise HoroshopPricesError("Вкажіть логін і пароль API або чинний токен.")
        response = self._post("/api/auth/", {"login": self.credentials.login, "password": self.credentials.password})
        token = response.get("response", {}).get("token")
        if not token: raise HoroshopPricesError("Хорошоп не повернув токен авторизації.")
        self._token = str(token)
        return self._token

    def export_catalog(self) -> list[dict[str, Any]]:
        offset, products = 0, []
        while True:
            response = self._post("/api/catalog/export/", {"token": self.token(), "offset": offset, "limit": DEFAULT_EXPORT_LIMIT, "includedParams": ["article_for_display", "price", "price_old", "wholesale_prices"]})
            nested = response.get("response")
            page = nested.get("products") if isinstance(nested, dict) else response.get("products")
            if not isinstance(page, list): raise HoroshopPricesError("Експорт каталогу не містить товарів.")
            products.extend(item for item in page if isinstance(item, dict))
            if len(page) < DEFAULT_EXPORT_LIMIT: return products
            offset += DEFAULT_EXPORT_LIMIT

    def import_products(self, products: list[dict[str, Any]]) -> dict[str, Any]:
        if not products: return {"status": "OK", "response": {"log": []}}
        return self._post("/api/catalog/import/", {"token": self.token(), "products": products})


def import_results(response: dict[str, Any]) -> dict[str, tuple[bool, str]]:
    status = str(response.get("status", "")).upper()
    nested = response.get("response")
    entries = nested.get("log", []) if isinstance(nested, dict) else []
    results: dict[str, tuple[bool, str]] = {}
    for entry in entries if isinstance(entries, list) else []:
        if not isinstance(entry, dict): continue
        article = normalize(entry.get("article"))
        info = entry.get("info", [])
        codes, messages = [], []
        for item in info if isinstance(info, list) else []:
            if isinstance(item, dict): codes.append(item.get("code")); messages.append(normalize(item.get("message")))
        results[article] = (0 in codes or (status == "OK" and not codes), "; ".join(message for message in messages if message))
    return results
